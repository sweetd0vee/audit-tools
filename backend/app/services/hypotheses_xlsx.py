"""Build and parse Excel checklist of audit hypotheses for a case."""

from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

COLUMNS: list[tuple[str, str, int]] = [
    ("n", "№", 5),
    ("hypothesis", "Гипотеза", 42),
    ("priority", "Приоритет", 12),
    ("risk", "Риск", 28),
    ("plan_sections", "Разделы плана", 24),
    ("npa_criteria", "НПА / критерии", 36),
    ("why_risk", "Почему это риск", 32),
    ("how_to_test", "Как проверить", 36),
    ("evidence_request", "Что запросить", 28),
    ("working_paper", "Рабочий документ", 24),
    ("basis", "Опора", 28),
]

_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "n": ("n", "№", "номер", "no", "num"),
    "hypothesis": ("hypothesis", "гипотеза", "формулировка", "title", "название"),
    "assertion": ("assertion", "утверждение"),
    "priority": ("priority", "приоритет"),
    "risk": ("risk", "риск"),
    "plan_sections": ("plan_sections", "разделы плана", "план", "программа"),
    "npa_criteria": ("npa_criteria", "нпа / критерии", "нпа", "критерии"),
    "why_risk": ("why_risk", "почему это риск", "почему риск"),
    "how_to_test": ("how_to_test", "как проверить", "проверка"),
    "evidence_request": ("evidence_request", "что запросить", "запрос"),
    "working_paper": ("working_paper", "рабочий документ", "рд"),
    "basis": ("basis", "опора"),
}

_SKIP_SHEETS = {"о проверке", "как читать"}
_SPACE_RE = re.compile(r"\s+")
_HEADER_STRIP_RE = re.compile(r"[«»\"'“”„#]")

_PRIORITY_ORDER = {"высокий": 0, "средний": 1, "низкий": 2}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
_HIGH_FILL = PatternFill("solid", fgColor="FCE4D6")
_MED_FILL = PatternFill("solid", fgColor="FFF2CC")
_WRAP = Alignment(wrap_text=True, vertical="top")
_THIN = Border(
    left=Side(style="thin", color="BDD7EE"),
    right=Side(style="thin", color="BDD7EE"),
    top=Side(style="thin", color="BDD7EE"),
    bottom=Side(style="thin", color="BDD7EE"),
)


def write_hypotheses_xlsx(
    path: Path,
    *,
    inspection_name: str,
    keywords: list[str],
    case_id: str,
    rows: list[dict[str, Any]],
    notes: str = "",
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Гипотезы"

    headers = [title for _, title, _ in COLUMNS]
    ws.append(headers)
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    keys = [key for key, _, _ in COLUMNS]
    ordered = sorted(
        rows,
        key=lambda r: _PRIORITY_ORDER.get(
            str(r.get("priority") or "").strip().lower(), 1
        ),
    )
    for i, row in enumerate(ordered, start=1):
        values: list[object] = []
        for key in keys:
            if key == "n":
                values.append(i)
            else:
                values.append(str(row.get(key) or "").strip())
        ws.append(values)
        excel_row = i + 1
        priority = str(row.get("priority") or "").strip().lower()
        fill = None
        if priority == "высокий":
            fill = _HIGH_FILL
        elif priority == "средний":
            fill = _MED_FILL
        for col_idx in range(1, len(keys) + 1):
            cell = ws.cell(excel_row, col_idx)
            cell.alignment = _WRAP
            cell.border = _THIN
            if fill is not None:
                cell.fill = fill
        ws.row_dimensions[excel_row].height = 60

    last_row = max(2, len(rows) + 1)
    last_col = get_column_letter(len(COLUMNS))
    table = Table(displayName="Hypotheses", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    meta = wb.create_sheet("О проверке", 1)
    meta["A1"] = "Чеклист гипотез внутренней аудиторской проверки"
    meta["A1"].font = Font(bold=True, size=14, color="1F4E79")
    meta["A3"] = "Проверка"
    meta["B3"] = inspection_name
    meta["A4"] = "Ключевые слова"
    meta["B4"] = ", ".join(keywords) or "—"
    meta["A5"] = "Кейс"
    meta["B5"] = case_id
    meta["A7"] = "Примечания модели"
    meta["B7"] = (notes or "").strip() or "—"
    meta["B7"].alignment = _WRAP
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 80
    meta.row_dimensions[7].height = 60

    legend = wb.create_sheet("Как читать", 2)
    legend["A1"] = "Колонки"
    legend["A1"].font = Font(bold=True)
    legend.append([])
    for _, title, _ in COLUMNS:
        legend.append([title])
    note_row = 3 + len(COLUMNS)
    legend[f"A{note_row}"] = (
        "Черновик для планирования СВА. Цитаты и номера статей сверять с файлами "
        "библиотеки кейса. Клиентские факты в этот контур ещё не входят. "
        "Строки отсортированы: высокий, средний, низкий приоритет. "
        "Свои гипотезы можно дописать в этот лист (колонка «Гипотеза») и "
        "приложить файл к команде «утверждаю гипотезы …»."
    )
    legend[f"A{note_row}"].alignment = _WRAP
    legend.column_dimensions["A"].width = 90
    legend.row_dimensions[note_row].height = 45

    wb.save(path)


def _norm_header(value: Any) -> str:
    text = _HEADER_STRIP_RE.sub("", str(value or "").strip().lower())
    return _SPACE_RE.sub(" ", text).strip(" .,:;")


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return _SPACE_RE.sub(" ", str(value).strip())


def _header_key(value: Any) -> str | None:
    header = _norm_header(value)
    if not header:
        return None
    for key, aliases in _HEADER_ALIASES.items():
        if header in aliases:
            return key
    return None


def _pick_sheet(wb) -> Any:
    for name in wb.sheetnames:
        if name.strip().lower() in _SKIP_SHEETS:
            continue
        if "гипотез" in name.strip().lower():
            return wb[name]
    for name in wb.sheetnames:
        if name.strip().lower() not in _SKIP_SHEETS:
            return wb[name]
    return wb.active


def _header_map(ws) -> tuple[int, dict[int, str]]:
    for row_idx in range(1, min(6, (ws.max_row or 1) + 1)):
        mapping: dict[int, str] = {}
        for col_idx in range(1, (ws.max_column or 1) + 1):
            key = _header_key(ws.cell(row_idx, col_idx).value)
            if key and key not in mapping.values():
                mapping[col_idx] = key
        if "hypothesis" in mapping.values():
            return row_idx, mapping
    raise ValueError(
        "В Excel нет колонки «Гипотеза». "
        "Скачайте чеклист `гипотезы` и допишите строки, либо сделайте лист "
        "с заголовком «Гипотеза»."
    )


def read_hypotheses_xlsx(data: bytes | Path) -> list[dict[str, str]]:
    """Parse auditor-supplied hypothesis rows from .xlsx bytes or a path."""
    if isinstance(data, Path):
        raw = data.read_bytes()
    else:
        raw = data
    if not raw:
        raise ValueError("Пустой Excel с гипотезами.")
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Не удалось открыть Excel с гипотезами: {exc}") from exc
    try:
        ws = _pick_sheet(wb)
        header_row, mapping = _header_map(ws)
        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            item: dict[str, str] = {}
            for cell in row:
                key = mapping.get(cell.column)
                if not key:
                    continue
                item[key] = _cell_text(cell.value)
            if item.get("hypothesis"):
                rows.append(item)
    finally:
        wb.close()
    if not rows:
        raise ValueError(
            "В Excel нет строк с гипотезами. Заполните колонку «Гипотеза»."
        )
    return rows
