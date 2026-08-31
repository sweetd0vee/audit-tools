import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.hypotheses_flow import parse_hypotheses_payload
from app.services.hypotheses_xlsx import read_hypotheses_xlsx, write_hypotheses_xlsx


def _row(i: int, priority: str) -> dict:
    return {
        "hypothesis": f"Гипотеза {i}",
        "assertion": "полнота",
        "risk": "недоплата",
        "plan_sections": "учёт аренды",
        "npa_criteria": "НК РБ [1]",
        "why_risk": "налоговый риск",
        "how_to_test": "сверка регистров",
        "evidence_request": "карточки счетов",
        "working_paper": "реестр отклонений",
        "priority": priority,
        "basis": "НПА [1]",
    }


class TestParseHypotheses(unittest.TestCase):
    def test_parses_object_with_eight_rows(self):
        payload = {
            "notes": "Черновик",
            "hypotheses": [_row(i, "высокий" if i <= 3 else "средний") for i in range(1, 9)],
        }
        rows, notes = parse_hypotheses_payload(payload)
        self.assertEqual(notes, "Черновик")
        self.assertEqual(len(rows), 8)
        self.assertEqual(rows[0]["priority"], "высокий")

    def test_sorts_high_then_medium_then_low(self):
        priorities = [
            "средний",
            "низкий",
            "высокий",
            "средний",
            "низкий",
            "высокий",
            "средний",
            "высокий",
        ]
        payload = {"hypotheses": [_row(i, p) for i, p in enumerate(priorities, start=1)]}
        rows, _ = parse_hypotheses_payload(payload)
        self.assertEqual([r["priority"] for r in rows], ["высокий"] * 3 + ["средний"] * 3 + ["низкий"] * 2)
        self.assertEqual([r["n"] for r in rows], [str(i) for i in range(1, 9)])
        self.assertEqual(rows[0]["hypothesis"], "Гипотеза 3")
        self.assertEqual(rows[3]["hypothesis"], "Гипотеза 1")
        self.assertEqual(rows[6]["hypothesis"], "Гипотеза 2")

    def test_rejects_too_few(self):
        with self.assertRaises(ValueError):
            parse_hypotheses_payload({"hypotheses": [{"hypothesis": "одна"}]})


class TestSelectHypotheses(unittest.TestCase):
    def test_resolve_keeps_order_and_dedupes(self):
        from app.services.hypotheses_flow import resolve_hypothesis_selection

        rows = [{**_row(i, "средний"), "n": str(i)} for i in range(1, 9)]
        self.assertEqual(
            resolve_hypothesis_selection(rows, numbers=[5, 1, 5, 3]),
            [5, 1, 3],
        )

    def test_auditor_origin_is_marked_in_prompt_block(self):
        from app.services.opinion_flow import format_hypotheses_block

        block = format_hypotheses_block(
            [{**_row(11, "высокий"), "n": "11", "origin": "auditor"}]
        )
        self.assertIn("[гипотеза аудитора]", block)


class TestHypothesesXlsx(unittest.TestCase):
    def test_writes_workbook_with_priority_column_and_sort(self):
        rows = [
            {
                "hypothesis": f"Гипотеза {i}",
                "assertion": "существование",
                "risk": "риск",
                "plan_sections": "раздел 7",
                "npa_criteria": "ГК РБ",
                "why_risk": "почему",
                "how_to_test": "как",
                "evidence_request": "договоры",
                "working_paper": "WP",
                "priority": ["низкий", "средний", "высокий"][i % 3],
                "basis": "программа",
            }
            for i in range(1, 9)
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "gipotezy.xlsx"
            write_hypotheses_xlsx(
                path,
                inspection_name="Проверка аренды",
                keywords=["аренда", "НДС"],
                case_id="abc123",
                rows=rows,
                notes="тест",
            )
            self.assertTrue(path.exists())
            self.assertGreater(path.stat().st_size, 2000)
            wb = load_workbook(path)
            ws = wb["Гипотезы"]
            headers = [cell.value for cell in ws[1]]
            self.assertEqual(headers[0], "№")
            self.assertEqual(headers[1], "Гипотеза")
            self.assertEqual(headers[2], "Приоритет")
            self.assertNotIn("Утверждение", headers)
            self.assertEqual(headers.count("Приоритет"), 1)
            priorities = [ws.cell(row, 3).value for row in range(2, ws.max_row + 1)]
            self.assertEqual(
                priorities,
                ["высокий"] * 3 + ["средний"] * 3 + ["низкий"] * 2,
            )
            self.assertEqual(ws.cell(2, 1).value, 1)
            self.assertEqual(ws.cell(2, 2).value, "Гипотеза 2")

    def test_reads_auditor_rows_and_skips_meta_sheets(self):
        rows = [_row(i, "средний") for i in range(1, 4)]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "gipotezy.xlsx"
            write_hypotheses_xlsx(
                path,
                inspection_name="Проверка аренды",
                keywords=["аренда"],
                case_id="abc123",
                rows=rows,
            )
            parsed = read_hypotheses_xlsx(path.read_bytes())
            self.assertEqual(len(parsed), 3)
            self.assertTrue(all(item.get("hypothesis") for item in parsed))

    def test_reads_simple_hypothesis_column(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист1"
        ws.append(["Гипотеза", "Приоритет"])
        ws.append(["Аренда без регистрации права", "высокий"])
        ws.append(["", "средний"])
        ws.append(["НДС принят к вычету без счёта-фактуры", "средний"])
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "mine.xlsx"
            wb.save(path)
            parsed = read_hypotheses_xlsx(path.read_bytes())
        self.assertEqual(
            [row["hypothesis"] for row in parsed],
            [
                "Аренда без регистрации права",
                "НДС принят к вычету без счёта-фактуры",
            ],
        )
        self.assertEqual(parsed[0]["priority"], "высокий")


class TestSelectWithAuditorExtras(unittest.TestCase):
    def setUp(self) -> None:
        from app.services.hypotheses_flow import _json_path
        from app.storage import store

        self.tmp = tempfile.TemporaryDirectory()
        self._prev = store.root
        store.root = Path(self.tmp.name)
        store.root.mkdir(parents=True, exist_ok=True)
        self.store = store
        self.state = store.create("Проверка аренды", ["аренда"])
        self.case_id = self.state.case_id
        rows = [{**_row(i, "высокий" if i <= 2 else "средний"), "n": str(i)} for i in range(1, 9)]
        path = _json_path(self.case_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps({"notes": "", "hypotheses": rows}, ensure_ascii=False),
            encoding="utf-8",
        )
        self.state.meta["hypotheses"] = {"built_at": "2026-01-01T00:00:00+00:00"}
        store.save(self.state)

    def tearDown(self) -> None:
        self.store.root = self._prev
        self.tmp.cleanup()

    def test_merges_json_extras_and_numbers_them_after_checklist(self):
        from app.services.hypotheses_flow import select_hypotheses, selected_hypothesis_rows

        result = select_hypotheses(
            self.case_id,
            numbers=[1, 2, 3, 4],
            extra_rows=[{"hypothesis": "Курсовые разницы не пересчитываются ежемесячно"}],
        )
        self.assertEqual(result["selected_ns"], [1, 2, 3, 4])
        self.assertEqual(result["extra_ns"], [9])
        self.assertEqual(result["extra_count"], 1)
        self.assertEqual(result["count"], 5)
        rows = selected_hypothesis_rows(self.store.get(self.case_id))
        self.assertEqual([r["n"] for r in rows], ["1", "2", "3", "4", "9"])
        self.assertEqual(rows[-1]["origin"], "auditor")
        self.assertIn("Курсовые разницы", rows[-1]["hypothesis"])

    def test_xlsx_skips_rows_already_in_checklist(self):
        from app.services.hypotheses_flow import select_hypotheses

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "extra.xlsx"
            write_hypotheses_xlsx(
                path,
                inspection_name="Проверка аренды",
                keywords=["аренда"],
                case_id=self.case_id,
                rows=[
                    _row(1, "высокий"),
                    {
                        **_row(99, "высокий"),
                        "hypothesis": "Договор аренды не зарегистрирован в установленном порядке",
                    },
                ],
            )
            result = select_hypotheses(
                self.case_id,
                numbers=[1, 2],
                extra_xlsx=path.read_bytes(),
                extra_filename="extra.xlsx",
            )
        self.assertEqual(result["extra_count"], 1)
        self.assertIn("не зарегистрирован", result["hypotheses"][-1]["hypothesis"])

    def test_add_only_keeps_previous_numbers(self):
        from app.services.hypotheses_flow import select_hypotheses, selected_hypothesis_rows

        select_hypotheses(self.case_id, numbers=[1, 3])
        select_hypotheses(
            self.case_id,
            keep_numbers=True,
            extra_rows=[{"hypothesis": "Лимит полномочий на подписание превышен"}],
        )
        rows = selected_hypothesis_rows(self.store.get(self.case_id))
        self.assertEqual([r["n"] for r in rows], ["1", "3", "9"])

    def test_api_json_and_multipart(self):
        from fastapi.testclient import TestClient

        from app.main import app

        client = TestClient(app)
        json_resp = client.post(
            f"/api/v1/cases/{self.case_id}/knowledge/hypotheses/select",
            json={
                "numbers": [1, 2],
                "extra_hypotheses": [{"hypothesis": "Своя гипотеза из JSON"}],
            },
        )
        self.assertEqual(json_resp.status_code, 200, json_resp.text)
        body = json_resp.json()
        self.assertEqual(body["selected_ns"], [1, 2])
        self.assertEqual(body["extra_count"], 1)

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "mine.xlsx"
            write_hypotheses_xlsx(
                path,
                inspection_name="Проверка аренды",
                keywords=["аренда"],
                case_id=self.case_id,
                rows=[
                    {
                        **_row(50, "средний"),
                        "hypothesis": "Залог оформлен без оценки предмета",
                    }
                ],
            )
            multi = client.post(
                f"/api/v1/cases/{self.case_id}/knowledge/hypotheses/select",
                data={"numbers": "[1, 4]", "keep_numbers": "false"},
                files={
                    "extra": (
                        "mine.xlsx",
                        path.read_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        self.assertEqual(multi.status_code, 200, multi.text)
        payload = multi.json()
        self.assertEqual(payload["selected_ns"], [1, 4])
        self.assertEqual(payload["extra_count"], 1)
        self.assertIn("Залог", payload["hypotheses"][-1]["hypothesis"])
        client.close()


if __name__ == "__main__":
    unittest.main()
